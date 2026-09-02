"""SustainSCM MRV Causal Scenario Completion Engine.

This module sits immediately before the existing SustainSCM KPI engine. It converts
partial scenario evidence into a complete, auditable seven-column MRV dataset.
It deliberately does not calculate normalized KPIs, composite indices, sensitivity
results, or MCDA rankings.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, Optional
import json
import math
import re
import uuid

import pandas as pd

UPLOAD_COLUMNS = [
    "variable_name",
    "value",
    "unit",
    "timestamp",
    "scenario_code",
    "source_system",
    "comment",
]

INPUT_SHEETS = {
    "scenario": "01_SCENARIO_CONFIG",
    "direct": "02_DIRECT_MRV_INPUT",
    "native": "03_NATIVE_MODEL_OUTPUTS",
    "assumptions": "04_APPROVED_ASSUMPTIONS",
    "reference": "05_REFERENCE_BASE",
}

OUTPUT_SHEETS = {
    "review": "11_COMPLETION_REVIEW",
    "upload": "12_SOFTWARE_UPLOAD",
    "qa": "13_QA_REPORT",
}

RULE_SOURCE_SYSTEM = {
    "L1": "direct_mrv_input",
    "L2": "mrv_derived_recalculation",
    "L3": "causal_baseline_scaling",
    "L4": "native_to_mrv_bridge",
    "L5": "approved_strategy_assumption",
    "L6": "causal_base_retention",
}


@dataclass(frozen=True)
class Permission:
    strategy_code: str
    variable_name: str
    influence_status: str
    permitted_rules: str
    priority: int
    justification: str
    resolution_source: str = "denied"

    def allows(self, level: str) -> bool:
        return level.upper() in _expand_rule_expression(self.permitted_rules)

    @property
    def blocks_change(self) -> bool:
        return self.influence_status == "Retain_BASE" or self.permitted_rules == "L6"


@dataclass
class CompletionResult:
    scenario_code: str
    run_id: str
    completion_review: pd.DataFrame
    software_upload: pd.DataFrame
    qa_report: pd.DataFrame
    rejected_inputs: pd.DataFrame
    l3_permission_diagnostics: pd.DataFrame
    rule_execution_trace: pd.DataFrame

    @property
    def has_critical_failures(self) -> bool:
        if self.qa_report.empty:
            return False
        return bool(
            ((self.qa_report["severity"] == "Critical") & (self.qa_report["status"] == "FAIL")).any()
        )

    def export_csv(self, path: str | Path, *, force: bool = False) -> Path:
        if self.has_critical_failures and not force:
            raise ValueError("Critical QA failures block CSV export. Review qa_report first.")
        target = Path(path)
        target.parent.mkdir(parents=True, exist_ok=True)
        self.software_upload.loc[:, UPLOAD_COLUMNS].to_csv(target, index=False)
        return target


class ScenarioCompletionEngine:
    """Complete scenario-level MRV measurements using a versioned L1-L6 protocol."""

    def __init__(
        self,
        config_dir: str | Path,
        *,
        config_frames: Mapping[str, pd.DataFrame] | None = None,
        engine_version: str = "1.1.0",
        rule_version: str = "1.1.0",
        strict_approval: bool = True,
        numerical_tolerance: float = 1e-6,
    ) -> None:
        self.config_dir = Path(config_dir)
        self.engine_version = engine_version
        self.rule_version = rule_version
        self.strict_approval = strict_approval
        self.numerical_tolerance = float(numerical_tolerance)

        supplied = config_frames or {}
        self.dictionary = _clean_frame(supplied.get("dictionary", self._read_config("mrv_dictionary.csv")))
        self.scope = _clean_frame(supplied.get("scope", self._read_config("strategy_scope.csv")))
        self.overrides = _clean_frame(supplied.get("overrides", self._read_config("variable_overrides.csv")))
        self.rules = _clean_frame(supplied.get("rules", self._read_config("mrv_rules.csv")))
        self.bridges = _clean_frame(supplied.get("bridges", self._read_config("bridge_rules.csv")))
        self.factor_register = _clean_frame(supplied.get("factor_register", pd.DataFrame()))
        self.default_factor_set_id = _text(supplied.get("default_factor_set_id"))

        self._validate_config()
        self.dictionary = self.dictionary.set_index("variable_name", drop=False)
        self._rule_order = self._topological_rule_order()
        self._permission_cache: dict[tuple[tuple[str, ...], str, str | None], Permission] = {}
        self._permission_details_cache: dict[tuple[tuple[str, ...], str], dict[str, Any]] = {}

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def complete_from_excel(
        self,
        workbook_path: str | Path,
        *,
        reference_csv: str | Path | None = None,
        output_csv: str | Path | None = None,
        write_back_workbook: str | Path | None = None,
        force_export: bool = False,
    ) -> CompletionResult:
        inputs = self.load_excel_inputs(workbook_path)
        if reference_csv is not None:
            reference = pd.read_csv(reference_csv)
        else:
            reference = inputs["reference"]

        result = self.complete(
            scenario=inputs["scenario"],
            direct_inputs=inputs["direct"],
            native_outputs=inputs["native"],
            assumptions=inputs["assumptions"],
            reference=reference,
        )

        if write_back_workbook is not None:
            self.write_outputs_to_workbook(
                source_workbook=workbook_path,
                destination_workbook=write_back_workbook,
                result=result,
            )
        if output_csv is not None:
            result.export_csv(output_csv, force=force_export)
        return result

    def load_excel_inputs(self, workbook_path: str | Path) -> dict[str, pd.DataFrame | pd.Series]:
        source = Path(workbook_path)
        if not source.exists():
            raise FileNotFoundError(source)
        xls = pd.ExcelFile(source)
        missing = [sheet for sheet in INPUT_SHEETS.values() if sheet not in xls.sheet_names]
        if missing:
            raise ValueError(f"Workbook is missing required sheets: {missing}")

        scenario_df = _clean_frame(pd.read_excel(source, sheet_name=INPUT_SHEETS["scenario"], header=3))
        if scenario_df.empty:
            raise ValueError("01_SCENARIO_CONFIG does not contain an active scenario row.")

        return {
            "scenario": scenario_df.iloc[0],
            "direct": _clean_frame(pd.read_excel(source, sheet_name=INPUT_SHEETS["direct"], header=3)),
            "native": _clean_frame(pd.read_excel(source, sheet_name=INPUT_SHEETS["native"], header=3)),
            "assumptions": _clean_frame(pd.read_excel(source, sheet_name=INPUT_SHEETS["assumptions"], header=3)),
            "reference": _clean_frame(pd.read_excel(source, sheet_name=INPUT_SHEETS["reference"], header=3)),
        }

    def complete(
        self,
        *,
        scenario: pd.Series | Mapping[str, Any],
        direct_inputs: pd.DataFrame,
        native_outputs: pd.DataFrame,
        assumptions: pd.DataFrame,
        reference: pd.DataFrame,
    ) -> CompletionResult:
        scenario = pd.Series(dict(scenario))
        scenario_code = _required_text(scenario.get("scenario_code"), "scenario_code")
        reference_code = _required_text(scenario.get("reference_scenario"), "reference_scenario")
        timestamp = _required_text(scenario.get("evaluation_timestamp"), "evaluation_timestamp")
        run_id = uuid.uuid4().hex[:16]
        selected_strategies = _selected_strategies(scenario)

        qa: list[dict[str, Any]] = []
        rejected: list[dict[str, Any]] = []
        l3_diagnostics: list[dict[str, Any]] = []
        rule_trace: list[dict[str, Any]] = []
        variable_issues: dict[str, list[tuple[str, str]]] = {}

        def issue(
            check_id: str,
            severity: str,
            status: str,
            message: str,
            *,
            variable: str = "",
            action: str = "Review the input or configuration.",
        ) -> None:
            qa.append(
                {
                    "check_id": check_id,
                    "severity": severity,
                    "status": status,
                    "affected_variable": variable,
                    "message": message,
                    "required_action": action,
                    "scenario_code": scenario_code,
                    "run_id": run_id,
                }
            )
            if variable and status in {"WARN", "FAIL"}:
                variable_issues.setdefault(variable, []).append((severity, message))

        if not selected_strategies:
            issue("QA_SCENARIO_STRATEGY", "Critical", "FAIL", "At least one strategy must be selected.")
        else:
            issue("QA_SCENARIO_STRATEGY", "Critical", "PASS", "At least one strategy is selected.")

        if _text(scenario.get("approval_status")) not in {"Approved", "Not required"}:
            issue(
                "QA_SCENARIO_APPROVAL",
                "Warning",
                "WARN",
                "The scenario configuration is not yet approved.",
                action="Approve the scenario before committing it to the database.",
            )
        else:
            issue("QA_SCENARIO_APPROVAL", "Warning", "PASS", "Scenario approval status is acceptable.")

        reference = self._prepare_reference(reference, reference_code)
        base_values = dict(zip(reference["variable_name"], reference["value"]))
        base_units = dict(zip(reference["variable_name"], reference["unit"]))
        base_comments = dict(zip(reference["variable_name"], reference["comment"]))
        common_variables = self.dictionary[
            self.dictionary["common_upload_variable"].map(_yes)
        ]["variable_name"].tolist()

        missing_reference = [v for v in common_variables if v not in base_values]
        if missing_reference:
            issue(
                "QA_REFERENCE_COMPLETENESS",
                "Critical",
                "FAIL",
                f"Reference scenario is missing {len(missing_reference)} common variables.",
                action="Load a complete validated BASE/reference dataset.",
            )
        else:
            issue("QA_REFERENCE_COMPLETENESS", "Critical", "PASS", "Reference scenario is complete.")

        direct_inputs = _clean_frame(direct_inputs)
        native_outputs = _clean_frame(native_outputs)
        assumptions = _clean_frame(assumptions)

        self._check_duplicates(direct_inputs, "variable_name", "direct input", issue)
        self._check_duplicates(native_outputs, "native_variable", "native output", issue)
        self._check_duplicates(assumptions, "assumption_id", "assumption", issue)

        # Start from the validated reference so every variable has a candidate,
        # but do not let that candidate conceal rejected competing evidence.
        # L6 is finalized only after all stronger evidence paths are evaluated.
        state: dict[str, dict[str, Any]] = {}
        for variable in common_variables:
            state[variable] = {
                "value": _float_or_none(base_values.get(variable)),
                "rule_level": "BASE" if scenario_code == reference_code else "L6",
                "rule_id": "BASE_REFERENCE" if scenario_code == reference_code else "BASE_RETENTION",
                "selected_strategy": "",
                "source_module": "Reference MRV",
                "source_reference": reference_code,
                "direct_input": None,
                "native_input": None,
                "provenance": (
                    f"Reference value retained from {reference_code}. "
                    f"{_text(base_comments.get(variable))}"
                ).strip(),
            }

        # Direct/common-MRV evidence --------------------------------------
        # Evidence class determines the rule level; a row's sheet location
        # does not turn translated or model-derived evidence into L1.
        direct_map: dict[str, float] = {}
        for _, row in direct_inputs.iterrows():
            variable = _text(row.get("variable_name"))
            value = _float_or_none(row.get("scenario_value"))
            if not variable or value is None:
                continue
            evidence_class = _text(
                row.get("normalized_evidence_class") or row.get("evidence_type")
            ).upper()
            evidence_level = {
                "DIRECT_MEASUREMENT": "L1",
                "DIRECT_MODEL_OUTPUT": "L1",
                "DERIVED_FROM_MODEL_OUTPUT": "L2",
                "CASE_SPECIFIC_BRIDGE": "L4",
                "APPROVED_ASSUMPTION": "L5",
                "BASE_RETENTION": "L6",
            }.get(evidence_class)
            if evidence_level is None:
                issue("QA_UNKNOWN_EVIDENCE_CLASS", "Critical", "FAIL", f"Unsupported evidence class: {evidence_class or 'blank'}.", variable=variable)
                rejected.append({"input_type": "direct", "variable_name": variable, "value": value, "reason": "Unsupported evidence class"})
                continue
            direct_map[variable] = value
            if variable not in self.dictionary.index:
                issue("QA_UNKNOWN_DIRECT_VARIABLE", "Critical", "FAIL", f"Unknown direct-input variable: {variable}", variable=variable)
                rejected.append({"input_type": "direct", "variable_name": variable, "value": value, "reason": "Unknown variable"})
                continue
            meta = self.dictionary.loc[variable]
            if evidence_level != "L2" and not _yes(meta.get("user_input_allowed")):
                issue("QA_DERIVED_DIRECT_INPUT", "Critical", "FAIL", "Direct input is not allowed for a derived/alias variable.", variable=variable)
                rejected.append({"input_type": "direct", "variable_name": variable, "value": value, "reason": "User input not allowed"})
                continue
            supplied_unit = _text(row.get("canonical_unit"))
            canonical_unit = _text(meta.get("canonical_unit"))
            if supplied_unit and canonical_unit and supplied_unit != canonical_unit:
                issue("QA_UNIT_MISMATCH", "Critical", "FAIL", f"Unit {supplied_unit!r} does not match canonical unit {canonical_unit!r}.", variable=variable)
                rejected.append({"input_type": "direct", "variable_name": variable, "value": value, "reason": "Unit mismatch"})
                continue
            if self.strict_approval and _text(row.get("approval_status")) != "Approved":
                issue("QA_UNAPPROVED_DIRECT_INPUT", "Warning", "WARN", "Direct input was ignored because it is not approved.", variable=variable, action="Approve the input or remove it.")
                rejected.append({"input_type": "direct", "variable_name": variable, "value": value, "reason": "Not approved"})
                continue
            permission = self.resolve_permission(selected_strategies, variable, evidence_level)
            if permission.blocks_change or not permission.allows(evidence_level):
                migrated_audit = _text(row.get("migration_disposition")) == "AUDIT_IF_OUTSIDE_CAUSAL_SCOPE"
                issue(
                    "QA_LEGACY_EVIDENCE_AUDIT_ONLY" if migrated_audit else "QA_UNAUTHORIZED_DIRECT_INPUT",
                    "Warning" if migrated_audit else "Critical",
                    "WARN" if migrated_audit else "FAIL",
                    (
                        f"Migrated legacy evidence for {variable} was preserved for audit but not selected as {evidence_level} because the declared strategies do not permit that causal change."
                        if migrated_audit else
                        f"Selected strategies do not permit an {evidence_level} change to {variable}."
                    ),
                    variable=variable,
                    action="Retain the completed BASE/derived value or configure a scientifically defensible strategy rule.",
                )
                rejected.append({
                    "input_type": "direct", "variable_name": variable, "value": value,
                    "reason": "Preserved for audit only" if migrated_audit else "Outside causal scope",
                })
                continue
            if evidence_level == "L6":
                # An explicit BASE-retention row is audit evidence, not a
                # competing scenario change.
                continue
            state[variable].update(
                value=value,
                rule_level=evidence_level,
                rule_id=evidence_class,
                selected_strategy=permission.strategy_code,
                source_module=_text(row.get("source_module")) or "Direct MRV",
                source_reference=_text(row.get("source_reference")),
                direct_input=value,
                provenance=f"Approved {evidence_class} evidence under {permission.strategy_code}.",
            )

        # L4 bridge transformations ---------------------------------------
        bridge_candidates: dict[str, list[dict[str, Any]]] = {}
        for _, row in native_outputs.iterrows():
            native_variable = _text(row.get("native_variable"))
            native_value = _float_or_none(row.get("scenario_value"))
            use = _yes(row.get("use_in_completion"))
            if not native_variable or native_value is None or not use:
                continue
            if self.strict_approval and _text(row.get("approval_status")) != "Approved":
                issue("QA_UNAPPROVED_NATIVE_OUTPUT", "Warning", "WARN", "Native output was ignored because it is not approved.", variable=native_variable)
                rejected.append({"input_type": "native", "variable_name": native_variable, "value": native_value, "reason": "Not approved"})
                continue
            rule_id = _text(row.get("bridge_rule_id"))
            target = _text(row.get("proposed_target_mrv"))
            if (
                rule_id == "BR_DES_TRANSPORT_GHG"
                and _text(row.get("model_family")).upper() != "DES"
            ):
                issue(
                    "INVALID_TRANSPORT_GHG_SOURCE", "Critical", "FAIL",
                    "A non-DES native output attempted to use the DES Vehicle-CO2 bridge.",
                    variable=native_variable,
                    action="Disable the placeholder or provide explicitly approved boundary-equivalent evidence.",
                )
                continue
            bridge = self._find_bridge(rule_id=rule_id, native_variable=native_variable, target_variable=target)
            if bridge is None:
                issue("QA_MISSING_BRIDGE", "Critical", "FAIL", "No active bridge rule was found for the native output.", variable=native_variable, action="Select or configure a documented bridge rule.")
                rejected.append({"input_type": "native", "variable_name": native_variable, "value": native_value, "reason": "Missing bridge"})
                continue
            target = _text(bridge["target_mrv_variable"])
            # Reference-scenario native rows are reconciliation/audit records;
            # the validated BASE reference remains authoritative.
            if scenario_code == reference_code:
                continue
            permission = self.resolve_permission(selected_strategies, target, "L4")
            if permission.blocks_change or not permission.allows("L4"):
                # An ACTIVE, approved, explicitly selected bridge is itself a
                # variable-level causal authorization. Group scope remains the
                # fallback for ordinary evidence and never authorizes a bridge
                # that is inactive, unapproved, or absent.
                permission = Permission(
                    _text(bridge.get("source_module")), target,
                    "Approved_active_bridge", "L4", 100,
                    _text(bridge.get("case_validity")),
                )
            try:
                transformed = self._evaluate_bridge(bridge, native_value, base_values)
            except Exception as exc:
                issue("QA_BRIDGE_ERROR", "Critical", "FAIL", f"Bridge {bridge['bridge_rule_id']} failed: {exc}", variable=target)
                continue
            bridge_candidates.setdefault(target, []).append(
                {
                    "value": transformed,
                    "native_value": native_value,
                    "bridge": bridge,
                    "permission": permission,
                    "source_reference": _text(row.get("source_reference")),
                }
            )

        for target, candidates in bridge_candidates.items():
            if len(candidates) > 1:
                unique = {round(float(c["value"]), 12) for c in candidates}
                if len(unique) > 1:
                    issue("QA_BRIDGE_CONFLICT", "Critical", "FAIL", "Multiple bridge rules produced conflicting values.", variable=target, action="Select one approved source or an integrated model result.")
                    continue
            candidate = candidates[0]
            if state[target]["rule_level"] == "L1":
                issue("QA_BRIDGE_SUPERSEDED", "Warning", "WARN", "Bridge result was superseded by a direct L1 input.", variable=target)
                continue
            bridge = candidate["bridge"]
            permission = candidate["permission"]
            state[target].update(
                value=candidate["value"],
                rule_level="L4",
                rule_id=_text(bridge["bridge_rule_id"]),
                selected_strategy=permission.strategy_code,
                source_module=_text(bridge["source_module"]),
                source_reference=candidate["source_reference"],
                native_input=candidate["native_value"],
                provenance=f"Native output transformed by bridge {_text(bridge['bridge_rule_id'])}.",
            )

        # L5 approved assumptions ----------------------------------------
        for _, row in assumptions.iterrows():
            status = _text(row.get("status"))
            variable = _text(row.get("target_variable"))
            value = _float_or_none(row.get("proposed_value"))
            if status != "Approved" or not variable or value is None:
                continue
            strategy = _text(row.get("strategy_code"))
            if strategy and strategy not in selected_strategies:
                issue("QA_UNUSED_ASSUMPTION", "Warning", "WARN", "Approved assumption was not used because its strategy is not selected.", variable=variable)
                continue
            permission = self.resolve_permission(selected_strategies, variable, "L5")
            if permission.blocks_change or not permission.allows("L5"):
                issue("QA_UNAUTHORIZED_ASSUMPTION", "Critical", "FAIL", "The approved assumption is outside the selected causal domain.", variable=variable)
                rejected.append({"input_type": "assumption", "variable_name": variable, "value": value, "reason": "Outside causal scope"})
                continue
            if state[variable]["rule_level"] in {"L1", "L4"}:
                issue("QA_ASSUMPTION_SUPERSEDED", "Warning", "WARN", "L5 assumption was superseded by stronger evidence.", variable=variable)
                continue
            state[variable].update(
                value=value,
                rule_level="L5",
                rule_id=_text(row.get("assumption_id")) or "APPROVED_ASSUMPTION",
                selected_strategy=permission.strategy_code,
                source_module="Approved assumption",
                source_reference=_text(row.get("evidence_or_expert_source")),
                provenance=f"Approved L5 assumption: {_text(row.get('justification'))}",
            )

        # L3 baseline-intensity scaling ----------------------------------
        scaling_enabled = _yes(scenario.get("allow_baseline_scaling"))
        driver = _text(scenario.get("scaling_driver_variable"))
        base_driver = _float_or_none(base_values.get(driver))
        scenario_driver = _float_or_none(state.get(driver, {}).get("value"))
        ratio = (
            scenario_driver / base_driver
            if scaling_enabled and base_driver not in (None, 0) and scenario_driver is not None
            else None
        )
        if scaling_enabled:
            if base_driver in (None, 0) or scenario_driver is None:
                issue("QA_SCALING_DRIVER", "Critical", "FAIL", "Baseline scaling was enabled but the scaling driver is unavailable or zero.", variable=driver)
        else:
            issue("QA_SCALING_SETTING", "Information", "PASS", "Baseline scaling is disabled.")

        for variable in common_variables:
            meta = self.dictionary.loc[variable]
            permission = self.resolve_permission(selected_strategies, variable, "L3")
            details = self.permission_details(selected_strategies, variable)
            explicitly_configured = permission.resolution_source == "exact override"
            dictionary_eligible = _is_activity_dependent_primary(meta)
            physical_activity_quantity = _is_physical_activity_quantity(meta)
            stronger_evidence = state[variable]["rule_level"] != "L6"
            eligible = bool(
                scaling_enabled and ratio is not None and not stronger_evidence
                and not permission.blocks_change
                and (permission.allows("L3") or (not explicitly_configured and physical_activity_quantity))
                and (explicitly_configured or dictionary_eligible or physical_activity_quantity)
            )
            reason = (
                "selected" if eligible else
                "baseline scaling disabled" if not scaling_enabled else
                "invalid or zero scaling driver" if ratio is None else
                f"stronger evidence {state[variable]['rule_level']}" if stronger_evidence else
                "L3 not permitted by exact override or strategy scope"
                if (not permission.allows("L3") and (explicitly_configured or not physical_activity_quantity))
                or permission.blocks_change else
                "not an explicitly configured or physical activity quantity"
            )
            if eligible:
                result_value = float(base_values[variable]) * float(ratio)
                state[variable].update(
                    value=result_value,
                    rule_level="L3",
                    rule_id="BASE_INTENSITY_SCALING",
                    selected_strategy=permission.strategy_code,
                    source_module="Baseline scaling",
                    source_reference=driver,
                    provenance=(
                        f"L3 scaling: BASE numerator={float(base_values[variable]):.12g}; "
                        f"BASE driver={base_driver:.12g}; scenario driver={scenario_driver:.12g}; "
                        f"ratio={ratio:.12g}; result={result_value:.12g}; "
                        f"permission_source={permission.resolution_source}; strategy={permission.strategy_code}."
                    ),
                )
            l3_diagnostics.append({
                "scenario_code": scenario_code, "variable_name": variable,
                "variable_group": _text(meta.get("variable_group")),
                "declared_strategies": ",".join(selected_strategies),
                **details,
                "final_permitted_rules": ",".join(sorted(_expand_rule_expression(permission.permitted_rules))),
                "dictionary_scaling_eligibility": dictionary_eligible,
                "physical_activity_quantity": physical_activity_quantity,
                "scaling_enabled": scaling_enabled, "driver_variable": driver,
                "driver_ratio": ratio, "L3_eligible": eligible,
                "L3_selected": eligible, "selection_reason": reason,
            })

        # Configured L3 identities (notably provisional coverage-preserving
        # volume rules) apply independently of general resource scaling.
        for _, rule in self.rules[self.rules["rule_level"].astype(str).str.upper() == "L3"].iterrows():
            target = _text(rule.get("target_variable"))
            if scenario_code == reference_code or target not in state:
                continue
            if state[target]["rule_level"] != "L6":
                continue
            permission = self.resolve_permission(selected_strategies, target, "L3")
            if permission.blocks_change or not permission.allows("L3"):
                continue
            try:
                value = self._evaluate_mrv_rule(rule, state, base_values, scenario=scenario)
            except Exception as exc:
                issue("QA_MRV_RULE_ERROR", "Critical", "FAIL", f"MRV rule {_text(rule['rule_id'])} failed: {exc}", variable=target)
                continue
            provenance = _text(rule.get("formula_description")) or "MRV identity recalculation."
            if _text(rule.get("operation")).upper() == "GHG_FROM_ENERGY_FACTORS":
                config = json.loads(_text(rule.get("parameter")))
                provenance = (
                    f"{provenance} factor_set_id={_text(scenario.get('emission_factor_set_id')) or self.default_factor_set_id}; "
                    f"electricity_factor_code={config['electricity_factor']}; diesel_factor_code={config['diesel_factor']}; "
                    f"electricity_kwh={state['electricity_kwh']['value']:.12g}; diesel_kwh={state['diesel_kwh']['value']:.12g}; "
                    f"result={value:.12g}."
                )
            state[target].update(
                value=value,
                rule_level="L3",
                rule_id=_text(rule.get("rule_id")),
                selected_strategy=permission.strategy_code,
                source_module="Provisional baseline scaling",
                source_reference=";".join(_rule_sources(rule)),
                provenance=_text(rule.get("formula_description")),
                provisional=True,
            )

        # L2 MRV identities and aliases in topological order -------------
        rules_by_target = self.rules.set_index("target_variable", drop=False)
        for target in self._rule_order:
            if scenario_code == reference_code:
                continue
            if target not in state or target not in rules_by_target.index:
                continue
            rule = rules_by_target.loc[target]
            if isinstance(rule, pd.DataFrame):
                rule = rule.iloc[0]
            if _text(rule.get("rule_level")) != "L2":
                continue
            sources = _rule_sources(rule)
            if state[target]["rule_level"] in {"L1", "L4", "L5"}:
                rule_trace.append({
                    "scenario_code": scenario_code, "rule_id": _text(rule.get("rule_id")),
                    "target_variable": target, "execution_order": len(rule_trace) + 1,
                    "source_variables": ",".join(sources), "source_values": "",
                    "factor_codes": "", "calculated_value": state[target]["value"],
                    "preserved_strong_evidence": True, "execution_status": "PRESERVED",
                    "message": f"Preserved {state[target]['rule_level']} evidence.",
                })
                continue
            # Documented model-equation evidence is already an L2 result. It
            # takes precedence over a configured fallback identity whose note
            # explicitly applies only when no scenario value is supplied.
            if (
                state[target]["rule_level"] == "L2"
                and state[target].get("rule_id") == "DERIVED_FROM_MODEL_OUTPUT"
            ):
                continue
            if (
                _text(rule.get("operation")).upper() == "GHG_FROM_ENERGY_FACTORS"
                and state[target]["rule_level"] == "L6"
                and all(state[source]["rule_level"] in {"BASE", "L6"} for source in sources)
            ):
                continue
            try:
                value = self._evaluate_mrv_rule(rule, state, base_values, scenario=scenario)
            except Exception as exc:
                issue("QA_MRV_RULE_ERROR", "Critical", "FAIL", f"MRV rule {_text(rule['rule_id'])} failed: {exc}", variable=target)
                rule_trace.append({
                    "scenario_code": scenario_code, "rule_id": _text(rule.get("rule_id")),
                    "target_variable": target, "execution_order": len(rule_trace) + 1,
                    "source_variables": ",".join(sources), "source_values": "",
                    "factor_codes": "", "calculated_value": None,
                    "preserved_strong_evidence": False, "execution_status": "FAILED",
                    "message": str(exc),
                })
                continue
            provenance = _text(rule.get("formula_description")) or "MRV identity recalculation."
            if _text(rule.get("operation")).upper() == "GHG_FROM_ENERGY_FACTORS":
                config = json.loads(_text(rule.get("parameter")))
                provenance = (
                    f"{provenance} factor_set_id={_text(scenario.get('emission_factor_set_id')) or self.default_factor_set_id}; "
                    f"electricity_factor_code={config['electricity_factor']}; diesel_factor_code={config['diesel_factor']}; "
                    f"electricity_kwh={state['electricity_kwh']['value']:.12g}; diesel_kwh={state['diesel_kwh']['value']:.12g}; "
                    f"result={value:.12g}."
                )
            elif _text(rule.get("operation")).upper() == "MULTIPLY_FACTOR":
                config = json.loads(_text(rule.get("parameter")))
                provenance = (
                    f"{provenance} factor_set_id={_text(scenario.get('emission_factor_set_id')) or self.default_factor_set_id}; "
                    f"factor_code={config['factor_code']}; source={state[sources[0]]['value']:.12g}; "
                    f"result={value:.12g}."
                )
            state[target].update(
                value=value,
                rule_level="L2",
                rule_id=_text(rule["rule_id"]),
                selected_strategy="",
                source_module="MRV calculation",
                source_reference=";".join(_rule_sources(rule)),
                provenance=provenance,
            )
            factor_codes = ""
            if _text(rule.get("operation")).upper() == "GHG_FROM_ENERGY_FACTORS":
                config = json.loads(_text(rule.get("parameter")))
                factor_codes = f"{config['electricity_factor']},{config['diesel_factor']}"
            elif _text(rule.get("operation")).upper() == "MULTIPLY_FACTOR":
                config = json.loads(_text(rule.get("parameter")))
                factor_codes = _text(config.get("factor_code"))
            rule_trace.append({
                "scenario_code": scenario_code, "rule_id": _text(rule.get("rule_id")),
                "target_variable": target, "execution_order": len(rule_trace) + 1,
                "source_variables": ",".join(sources),
                "source_values": ",".join(f"{name}={state[name]['value']}" for name in sources if name in state),
                "factor_codes": factor_codes, "calculated_value": value,
                "preserved_strong_evidence": False, "execution_status": "EXECUTED",
                "message": "Rule executed after final primary-value resolution.",
            })

        # A direct customer-acceptance index remains authoritative, but its
        # consistency with the canonical equal-weight formula is auditable.
        if "customer_acceptance_index" in direct_map:
            components = [
                _float_or_none(state.get(name, {}).get("value"))
                for name in (
                    "sustainable_sales_share",
                    "customer_survey_score",
                    "contract_renewal_rate",
                )
            ]
            if all(value is not None for value in components):
                calculated = sum(components) / 3.0 * 100.0
                difference = abs(direct_map["customer_acceptance_index"] - calculated)
                if difference > self.numerical_tolerance:
                    issue(
                        "QA_CUSTOMER_ACCEPTANCE_CONFLICT",
                        "Warning",
                        "WARN",
                        (
                            "Direct customer acceptance differs from the canonical "
                            f"equal-weight calculation by {difference:.12g} points; "
                            "the approved L1 value was retained."
                        ),
                        variable="customer_acceptance_index",
                    )

        # Blocking stale-value checks -----------------------------------
        for diagnostic in l3_diagnostics:
            variable = diagnostic["variable_name"]
            if diagnostic["L3_eligible"] and state[variable]["rule_level"] == "L6":
                issue(
                    "DES_ACTIVITY_VALUE_RETAINED_AT_BASE", "Critical", "FAIL",
                    "L3 was permitted and eligible, but the activity value retained BASE through L6.",
                    variable=variable,
                    action="Apply configured BASE-intensity scaling before L6 retention.",
                )
        energy_changed = any(
            not math.isclose(float(state[name]["value"]), float(base_values[name]), rel_tol=0.0, abs_tol=self.numerical_tolerance)
            for name in ("electricity_kwh", "diesel_kwh")
            if name in state and name in base_values
        )
        ghg = state.get("ghg_total_s1s2", {})
        if energy_changed and (
            ghg.get("rule_level") == "L6"
            or (
                ghg.get("rule_level") not in {"L1", "L4", "L5"}
                and math.isclose(
                    float(ghg.get("value")), float(base_values["ghg_total_s1s2"]),
                    rel_tol=0.0, abs_tol=self.numerical_tolerance,
                )
            )
        ):
            issue(
                "DERIVED_GHG_STALE", "Critical", "FAIL",
                "Energy changed but total Scope 1-2 GHG retained the reference result.",
                variable="ghg_total_s1s2",
                action="Execute the configured factor-dependent GHG rule after L3.",
            )
        intensity_dependencies = {
            "energy_intensity_fu": ("total_energy_kwh", "output_qty_fu"),
            "waste_generation_intensity_fu": ("waste_generated_t", "output_qty_fu"),
            "water_intensity_fu": ("water_withdrawn_m3", "output_qty_fu"),
            "cost_per_fu": ("operating_cost_eur", "output_qty_fu"),
            "ghg_intensity_fu": ("ghg_total_s1s2", "output_qty_fu"),
            "transport_ghg_intensity": ("transport_ghg_tco2e", "transport_work_tkm"),
        }
        for target, dependencies in intensity_dependencies.items():
            if target not in state or any(name not in state for name in dependencies):
                continue
            dependency_changed = any(
                not math.isclose(float(state[name]["value"]), float(base_values[name]), rel_tol=0.0, abs_tol=self.numerical_tolerance)
                for name in dependencies
            )
            if dependency_changed and state[target]["rule_level"] == "L6":
                issue(
                    "DERIVED_INTENSITY_STALE", "Critical", "FAIL",
                    "An intensity retained BASE provenance after its numerator or denominator changed.",
                    variable=target,
                    action="Recalculate the derived intensity after final primary-value resolution.",
                )

        # QA bounds and cross-variable consistency -----------------------
        for variable, record in state.items():
            value = _float_or_none(record.get("value"))
            if value is None or not math.isfinite(value):
                issue("QA_MISSING_COMPLETED_VALUE", "Critical", "FAIL", "No finite completed value is available.", variable=variable)
                continue
            meta = self.dictionary.loc[variable]
            min_value = _float_or_none(meta.get("min_value"))
            max_value = _float_or_none(meta.get("max_value"))
            if min_value is not None and value < min_value:
                issue("QA_MIN_BOUND", "Critical", "FAIL", f"Value {value} is below minimum {min_value}.", variable=variable)
            provisional_dpp_coverage = (
                variable == "dpp_coverage"
                and state.get("dpp_valid_volume", {}).get("rule_level") == "L6"
            )
            if max_value is not None and value > max_value and not provisional_dpp_coverage:
                issue("QA_MAX_BOUND", "Critical", "FAIL", f"Value {value} is above maximum {max_value}.", variable=variable)

        self._relationship_check(state, "renewable_energy_kwh", "electricity_kwh", "Renewable energy exceeds total electricity.", issue)
        self._relationship_check(state, "waste_recovered_t", "waste_generated_t", "Recovered waste exceeds waste generated.", issue)
        self._relationship_check(state, "material_circular_t", "material_total_t", "Circular material exceeds total material.", issue)
        dpp_valid = state.get("dpp_valid_volume", {})
        shipped = state.get("shipped_volume_total", {})
        retained_dpp_valid = _float_or_none(base_values.get("dpp_valid_volume"))
        shipped_value = _float_or_none(shipped.get("value"))
        # A MIN accounting identity can bound a preliminary value, but cannot
        # prove batch/event validity. Keep the pre-DPP fallback explicit until
        # the DPP service supplies authoritative evidence in pass 2.
        if (
            retained_dpp_valid is not None
            and shipped_value is not None
            and retained_dpp_valid > shipped_value + self.numerical_tolerance
        ):
            issue(
                "QA_DPP_VALIDATION_PENDING",
                "Warning",
                "WARN",
                "Reference DPP-valid volume exceeds scenario shipped volume; the bounded L2 value remains provisional until validated batch/event evidence overwrites it.",
                variable="dpp_valid_volume",
                action="Import validated DPP batches for this scenario before final KPI publication.",
            )
        if (
            _float_or_none(dpp_valid.get("value")) is not None
            and _float_or_none(shipped.get("value")) is not None
            and _float_or_none(dpp_valid.get("value"))
            > _float_or_none(shipped.get("value")) + self.numerical_tolerance
        ):
            if dpp_valid.get("rule_level") == "L6":
                issue(
                    "QA_DPP_VALIDATION_PENDING",
                    "Warning",
                    "WARN",
                    "The retained reference DPP-valid volume exceeds scenario shipped volume; replace it with validated DPP batch evidence before interpreting DPP coverage.",
                    variable="dpp_valid_volume",
                    action="Import validated DPP batches for this scenario; the provisional L6 value does not block MRV scenario completion.",
                )
            else:
                issue("QA_RELATIONSHIP", "Critical", "FAIL", "DPP-valid volume exceeds shipped volume.", variable="dpp_valid_volume")
        self._relationship_check(state, "mrv_points_active_valid", "mrv_points_required", "Valid MRV points exceed required MRV points.", issue)

        energy_changed = any(
            abs(float(state[name]["value"]) - float(base_values[name])) > self.numerical_tolerance
            for name in ("electricity_kwh", "diesel_kwh")
            if name in state and name in base_values
        )
        if energy_changed and state.get("ghg_total_s1s2", {}).get("rule_level") == "L6":
            issue(
                "QA_STALE_DERIVED_GHG",
                "Critical",
                "FAIL",
                "Electricity or diesel differs from BASE, but total Scope 1-2 GHG was retained through L6.",
                variable="ghg_total_s1s2",
                action="Execute an approved factor-dependent L2 rule after final energy resolution.",
            )

        transport = state.get("transport_ghg_tco2e", {})
        transport_rule_ids = set(
            self.rules.loc[
                (self.rules["target_variable"].astype(str).str.strip() == "transport_ghg_tco2e")
                & (self.rules["operation"].astype(str).str.strip().str.upper() == "MULTIPLY_FACTOR"),
                "rule_id",
            ].astype(str)
        )
        transport_provenance = _text(transport.get("provenance")).lower()
        boundary_ok = (
            transport.get("rule_level") == "BASE"
            or transport.get("rule_id") in transport_rule_ids
            or transport.get("rule_id") == "BR_DES_TRANSPORT_GHG"
        ) and not any(token in transport_provenance for token in ("ef_diesel", "scope 1 diesel"))
        issue(
            "TRANSPORT_GHG_BOUNDARY_MISMATCH",
            "Critical",
            "PASS" if boundary_ok else "FAIL",
            (
                "Transport GHG uses the approved outbound-road transport boundary."
                if boundary_ok else
                "Transport GHG does not use an approved transport-work factor or DES vehicle-CO2 bridge."
            ),
            variable="transport_ghg_tco2e",
            action="Use transport_work_tkm × the approved tCO2e/tkm factor, or genuine DES Vehicle CO2 evidence.",
        )
        active_rules = self.rules[
            self.rules.get("rule_status", pd.Series("ACTIVE", index=self.rules.index)).astype(str).str.upper().eq("ACTIVE")
        ]
        double_count_rules = active_rules[
            (active_rules["target_variable"].astype(str).str.strip() == "ghg_total_s1s2")
            & active_rules.apply(lambda row: "transport_ghg_tco2e" in _rule_sources(row), axis=1)
        ]
        issue(
            "TRANSPORT_GHG_DOUBLE_COUNT_RISK",
            "Critical",
            "PASS" if double_count_rules.empty else "FAIL",
            (
                "Scope 1+2 GHG is independent of the outbound transport-GHG numerator."
                if double_count_rules.empty else
                "An active Scope 1+2 rule includes transport GHG and may double count E7 emissions."
            ),
            variable="ghg_total_s1s2",
            action="Remove transport_ghg_tco2e from active Scope 1+2 calculation rules.",
        )

        # Build outputs ---------------------------------------------------
        review_rows: list[dict[str, Any]] = []
        upload_rows: list[dict[str, Any]] = []
        for variable in common_variables:
            record = state[variable]
            issues = variable_issues.get(variable, [])
            qa_status = "FAIL" if any(sev == "Critical" for sev, _ in issues) else "WARN" if issues else "PASS"
            qa_message = " | ".join(message for _, message in issues)
            rule_level = record["rule_level"]
            source_system = self._source_system(rule_level, record.get("source_module"))
            comment = (
                f"rule={rule_level}; rule_id={record['rule_id']}; "
                f"strategy={record.get('selected_strategy') or 'none'}; "
                f"provenance={record.get('provenance', '')}"
            )
            review_rows.append(
                {
                    "variable_name": variable,
                    "unit": base_units.get(variable, _text(self.dictionary.loc[variable, "canonical_unit"])),
                    "base_value": _float_or_none(base_values.get(variable)),
                    "direct_input": record.get("direct_input"),
                    "native_input": record.get("native_input"),
                    "completed_value": record.get("value"),
                    "rule_level": rule_level,
                    "rule_id": record.get("rule_id"),
                    "selected_strategy": record.get("selected_strategy"),
                    "source_module": record.get("source_module"),
                    "source_reference": record.get("source_reference"),
                    "provenance": record.get("provenance"),
                    "provisional": bool(record.get("provisional", False)),
                    "qa_status": qa_status,
                    "qa_message": qa_message,
                    "timestamp": timestamp,
                    "scenario_code": scenario_code,
                    "engine_version": self.engine_version,
                    "rule_version": self.rule_version,
                }
            )
            upload_rows.append(
                {
                    "variable_name": variable,
                    "value": record.get("value"),
                    "unit": base_units.get(variable, _text(self.dictionary.loc[variable, "canonical_unit"])),
                    "timestamp": timestamp,
                    "scenario_code": scenario_code,
                    "source_system": source_system,
                    "comment": comment,
                }
            )

        completed_count = sum(_float_or_none(r["value"]) is not None for r in upload_rows)
        if completed_count == len(common_variables):
            issue("QA_COMPLETION_COUNT", "Critical", "PASS", f"All {len(common_variables)} common variables were completed.")
        else:
            issue("QA_COMPLETION_COUNT", "Critical", "FAIL", f"Only {completed_count} of {len(common_variables)} common variables were completed.")

        qa_df = pd.DataFrame(qa, columns=[
            "check_id", "severity", "status", "affected_variable", "message", "required_action", "scenario_code", "run_id"
        ])
        review_df = pd.DataFrame(review_rows)
        upload_df = pd.DataFrame(upload_rows, columns=UPLOAD_COLUMNS)
        rejected_df = pd.DataFrame(rejected, columns=["input_type", "variable_name", "value", "reason"])

        return CompletionResult(
            scenario_code=scenario_code,
            run_id=run_id,
            completion_review=review_df,
            software_upload=upload_df,
            qa_report=qa_df,
            rejected_inputs=rejected_df,
            l3_permission_diagnostics=pd.DataFrame(l3_diagnostics),
            rule_execution_trace=pd.DataFrame(rule_trace),
        )

    def permission_details(
        self, selected_strategies: Iterable[str], variable_name: str
    ) -> dict[str, Any]:
        """Return auditable exact/scope inputs without changing resolution."""
        strategies = tuple(selected_strategies)
        cache_key = (strategies, variable_name)
        if cache_key in self._permission_details_cache:
            return dict(self._permission_details_cache[cache_key])
        group = _text(self.dictionary.loc[variable_name, "variable_group"])
        exact_rows: list[str] = []
        active_rows: list[str] = []
        scope_rows: list[str] = []
        for strategy in strategies:
            exact = self.overrides[
                (self.overrides["strategy_code"] == strategy)
                & (self.overrides["variable_name"] == variable_name)
            ]
            for _, row in exact.iterrows():
                rendered = f"{strategy}:{_text(row.get('permitted_rules'))}"
                exact_rows.append(rendered)
                if _yes(row.get("active")):
                    active_rows.append(rendered)
            scope = self.scope[
                (self.scope["strategy_code"] == strategy)
                & (self.scope["variable_group"] == group)
            ]
            for _, row in scope.iterrows():
                scope_rows.append(f"{strategy}:{_text(row.get('permitted_rules'))}")
        result = {
            "exact_override_found": bool(exact_rows),
            "exact_override_active": bool(active_rows),
            "exact_override_rules": ";".join(exact_rows),
            "scope_rules": ";".join(scope_rows),
        }
        self._permission_details_cache[cache_key] = result
        return dict(result)

    def resolve_permission(
        self,
        selected_strategies: Iterable[str],
        variable_name: str,
        rule_level: str | None = None,
    ) -> Permission:
        strategies = tuple(selected_strategies)
        cache_key = (strategies, variable_name, rule_level)
        if cache_key not in self._permission_cache:
            self._permission_cache[cache_key] = self._resolve_permission_uncached(
                strategies, variable_name, rule_level
            )
        return self._permission_cache[cache_key]

    def _resolve_permission_uncached(
        self,
        selected_strategies: Iterable[str],
        variable_name: str,
        rule_level: str | None = None,
    ) -> Permission:
        if variable_name not in self.dictionary.index:
            return Permission("", variable_name, "Retain_BASE", "L6", 0, "Unknown variable.")
        group = _text(self.dictionary.loc[variable_name, "variable_group"])
        candidates: list[Permission] = []
        exact_candidates: list[Permission] = []
        for strategy in selected_strategies:
            if not strategy:
                continue
            exact = self.overrides[
                (self.overrides["strategy_code"] == strategy)
                & (self.overrides["active"].map(_yes))
                & (self.overrides["variable_name"].isin([variable_name, "*"]))
            ]
            if not exact.empty:
                # Exact variable row wins over wildcard row.
                exact = exact.assign(_specific=(exact["variable_name"] == variable_name).astype(int))
                row = exact.sort_values(["_specific", "priority"], ascending=False).iloc[0]
                exact_candidates.append(
                    Permission(
                        strategy,
                        variable_name,
                        _text(row["influence_status"]),
                        _text(row["permitted_rules"]),
                        int(row["priority"]),
                        _text(row["scientific_justification"]),
                        "exact override",
                    )
                )
                continue
            scope = self.scope[
                (self.scope["strategy_code"] == strategy)
                & (self.scope["variable_group"] == group)
            ]
            if not scope.empty:
                row = scope.sort_values("priority", ascending=False).iloc[0]
                candidates.append(
                    Permission(
                        strategy,
                        variable_name,
                        _text(row["influence_status"]),
                        _text(row["permitted_rules"]),
                        int(row["priority"]),
                        _text(row["scientific_interpretation"]),
                        "strategy scope",
                    )
                )
        # Exact rows are exceptional declarations, not a whitelist.  Resolve
        # contradictory exact declarations by priority, while strategies with
        # no exact row retain their independently resolved group permission.
        if exact_candidates:
            candidates.append(sorted(exact_candidates, key=lambda p: p.priority, reverse=True)[0])
        if rule_level:
            permitting = [
                candidate for candidate in candidates
                if not candidate.blocks_change and candidate.allows(rule_level)
            ]
            if permitting:
                if rule_level.upper() == "L3":
                    scaling = [
                        candidate for candidate in permitting
                        if "scaled" in candidate.influence_status.lower()
                        or "scaling" in candidate.influence_status.lower()
                    ]
                    if scaling:
                        permitting = scaling
                return sorted(permitting, key=lambda p: p.priority, reverse=True)[0]
        if not candidates:
            return Permission("", variable_name, "Retain_BASE", "L6", 0, "No selected strategy supplied a causal permission.")
        return sorted(candidates, key=lambda p: p.priority, reverse=True)[0]

    def write_outputs_to_workbook(
        self,
        *,
        source_workbook: str | Path,
        destination_workbook: str | Path,
        result: CompletionResult,
    ) -> Path:
        """Write review, upload, and QA tables while preserving the template layout."""
        from copy import copy
        from openpyxl import load_workbook

        source = Path(source_workbook)
        destination = Path(destination_workbook)
        destination.parent.mkdir(parents=True, exist_ok=True)
        wb = load_workbook(source)

        self._write_dataframe_to_sheet(wb[OUTPUT_SHEETS["review"]], result.completion_review)
        self._write_dataframe_to_sheet(wb[OUTPUT_SHEETS["upload"]], result.software_upload)
        self._write_dataframe_to_sheet(wb[OUTPUT_SHEETS["qa"]], result.qa_report)
        wb.save(destination)
        return destination

    @staticmethod
    def _write_dataframe_to_sheet(worksheet: Any, dataframe: pd.DataFrame) -> None:
        from copy import copy

        header_row = 4
        start_row = 5
        headers = [worksheet.cell(header_row, col).value for col in range(1, worksheet.max_column + 1)]
        expected = [str(h) for h in headers if h is not None]
        missing = [c for c in expected if c not in dataframe.columns]
        if missing:
            raise ValueError(f"Output DataFrame does not contain workbook columns: {missing}")

        # Clear old values but keep styles.
        for row in worksheet.iter_rows(min_row=start_row, max_row=max(worksheet.max_row, start_row)):
            for cell in row:
                cell.value = None

        template_row = start_row
        for ridx, record in enumerate(dataframe.loc[:, expected].itertuples(index=False, name=None), start=start_row):
            if ridx > worksheet.max_row:
                worksheet.insert_rows(ridx)
            for cidx, value in enumerate(record, start=1):
                cell = worksheet.cell(ridx, cidx)
                if ridx != template_row:
                    source_cell = worksheet.cell(template_row, cidx)
                    if source_cell.has_style:
                        cell._style = copy(source_cell._style)
                    if source_cell.number_format:
                        cell.number_format = source_cell.number_format
                    cell.alignment = copy(source_cell.alignment)
                    cell.fill = copy(source_cell.fill)
                    cell.font = copy(source_cell.font)
                    cell.border = copy(source_cell.border)
                cell.value = None if pd.isna(value) else value

    # ------------------------------------------------------------------
    # Configuration and calculations
    # ------------------------------------------------------------------
    def _read_config(self, filename: str) -> pd.DataFrame:
        path = self.config_dir / filename
        if not path.exists():
            raise FileNotFoundError(f"Missing engine configuration: {path}")
        return _clean_frame(pd.read_csv(path))

    def _validate_config(self) -> None:
        required = {
            "dictionary": {"variable_name", "canonical_unit", "variable_group", "common_upload_variable", "user_input_allowed", "min_value", "max_value"},
            "scope": {"strategy_code", "variable_group", "influence_status", "permitted_rules", "priority"},
            "overrides": {"strategy_code", "variable_name", "influence_status", "permitted_rules", "priority", "active"},
            "rules": {"rule_id", "target_variable", "operation", "source_1", "source_2", "source_3", "parameter", "rule_status"},
            "bridges": {"bridge_rule_id", "source_module", "native_variable", "target_mrv_variable", "operation", "parameter", "base_multiplier_variable", "rule_status"},
        }
        frames = {
            "dictionary": self.dictionary,
            "scope": self.scope,
            "overrides": self.overrides,
            "rules": self.rules,
            "bridges": self.bridges,
        }
        for name, columns in required.items():
            missing = columns - set(frames[name].columns)
            if missing:
                raise ValueError(f"{name} configuration is missing columns: {sorted(missing)}")
        if self.dictionary["variable_name"].duplicated().any():
            raise ValueError("mrv_dictionary.csv contains duplicate variable_name values.")
        for frame_name, frame in (("strategy scope", self.scope), ("variable overrides", self.overrides)):
            for value in frame.get("permitted_rules", pd.Series(dtype=object)).dropna().unique():
                try:
                    parse_completion_levels(value)
                except ValueError as exc:
                    raise ValueError(f"Malformed {frame_name} permitted_rules {value!r}: {exc}") from exc

    def _topological_rule_order(self) -> list[str]:
        active = self.rules[self.rules["rule_status"].astype(str).str.upper() == "ACTIVE"]
        targets = set(active["target_variable"])
        dependencies: dict[str, set[str]] = {}
        for _, row in active.iterrows():
            target = _text(row["target_variable"])
            dependencies[target] = {src for src in _rule_sources(row) if src in targets}
        order: list[str] = []
        remaining = {k: set(v) for k, v in dependencies.items()}
        while remaining:
            ready = sorted([target for target, deps in remaining.items() if not deps])
            if not ready:
                raise ValueError(f"Circular MRV-rule dependency detected: {remaining}")
            for target in ready:
                order.append(target)
                remaining.pop(target)
                for deps in remaining.values():
                    deps.discard(target)
        return order

    def _prepare_reference(self, reference: pd.DataFrame, reference_code: str) -> pd.DataFrame:
        required = set(UPLOAD_COLUMNS)
        missing = required - set(reference.columns)
        if missing:
            raise ValueError(f"Reference data is missing required CSV columns: {sorted(missing)}")
        ref = reference.copy()
        if "scenario_code" in ref.columns:
            filtered = ref[ref["scenario_code"].astype(str) == reference_code]
            if not filtered.empty:
                ref = filtered
        ref = ref.loc[:, UPLOAD_COLUMNS].copy()
        ref["value"] = pd.to_numeric(ref["value"], errors="coerce")
        ref = ref.dropna(subset=["variable_name", "value"])
        if ref["variable_name"].duplicated().any():
            duplicates = ref.loc[ref["variable_name"].duplicated(), "variable_name"].tolist()
            raise ValueError(f"Reference scenario contains duplicate variables: {duplicates}")
        return ref

    @staticmethod
    def _check_duplicates(frame: pd.DataFrame, key: str, label: str, issue: Callable[..., None]) -> None:
        if key not in frame.columns or frame.empty:
            return
        populated = frame[frame[key].notna() & (frame[key].astype(str).str.strip() != "")]
        duplicates = populated[populated[key].duplicated(keep=False)][key].astype(str).unique().tolist()
        if duplicates:
            issue("QA_DUPLICATE_INPUT", "Critical", "FAIL", f"Duplicate {label} keys: {duplicates}")
        else:
            issue(f"QA_DUPLICATE_{key.upper()}", "Critical", "PASS", f"No duplicate {label} keys were found.")

    def _find_bridge(self, *, rule_id: str, native_variable: str, target_variable: str) -> Optional[pd.Series]:
        active = self.bridges[self.bridges["rule_status"].astype(str).str.upper() == "ACTIVE"]
        if rule_id:
            match = active[active["bridge_rule_id"] == rule_id]
        else:
            match = active[active["native_variable"] == native_variable]
            if target_variable:
                match = match[match["target_mrv_variable"] == target_variable]
        if match.empty:
            return None
        if len(match) > 1 and not target_variable:
            return None
        return match.iloc[0]

    @staticmethod
    def _evaluate_bridge(bridge: pd.Series, native_value: float, base_values: Mapping[str, float]) -> float:
        operation = _text(bridge["operation"]).upper()
        parameter = _float_or_none(bridge.get("parameter"))
        if operation == "ALIAS":
            return float(native_value)
        if operation == "MULTIPLY_CONSTANT":
            if parameter is None:
                raise ValueError("MULTIPLY_CONSTANT requires parameter.")
            return float(native_value) * parameter
        if operation == "DIVIDE_CONSTANT":
            if parameter in (None, 0):
                raise ValueError("DIVIDE_CONSTANT requires a nonzero parameter.")
            return float(native_value) / parameter
        if operation == "MULTIPLY_BASE":
            base_variable = _text(bridge.get("base_multiplier_variable"))
            base_value = _float_or_none(base_values.get(base_variable))
            if base_value is None:
                raise ValueError(f"BASE multiplier variable {base_variable!r} is unavailable.")
            return float(native_value) * base_value
        raise ValueError(f"Unsupported bridge operation: {operation}")

    def _evaluate_mrv_rule(
        self,
        rule: pd.Series,
        state: Mapping[str, Mapping[str, Any]],
        base_values: Mapping[str, float],
        *,
        scenario: Mapping[str, Any] | None = None,
    ) -> float:
        operation = _text(rule["operation"]).upper()
        sources = _rule_sources(rule)
        values = []
        for source in sources:
            value = _float_or_none(state.get(source, {}).get("value"))
            if value is None:
                raise ValueError(f"Source variable {source!r} is unavailable.")
            values.append(value)
        parameter = _float_or_none(rule.get("parameter"))
        factor_dispatch = {
            "GHG_FROM_ENERGY_FACTORS": self._ghg_from_energy_factors,
            "MULTIPLY_FACTOR": self._multiply_factor,
        }
        if operation in factor_dispatch:
            try:
                config = json.loads(_text(rule.get("parameter")))
            except (TypeError, json.JSONDecodeError) as exc:
                raise ValueError(f"{operation} requires a valid JSON parameter object") from exc
            if not isinstance(config, dict):
                raise ValueError(f"{operation} parameter must be a JSON object")
            return factor_dispatch[operation](values, config, {} if scenario is None else scenario)
        if operation == "ALIAS":
            return values[0]
        if operation == "SUM":
            return sum(values)
        if operation == "PRODUCT":
            return values[0] * values[1]
        if operation == "RATIO":
            return _safe_divide(values[0], values[1])
        if operation == "RATIO_PCT":
            return _safe_divide(values[0], values[1]) * 100.0
        if operation == "RATIO_X1000":
            return _safe_divide(values[0] * 1000.0, values[1])
        if operation == "RATIO_X1M":
            return _safe_divide(values[0] * 1_000_000.0, values[1])
        if operation == "RATIO_DIV24":
            return _safe_divide(values[0], values[1]) / 24.0
        if operation == "AVERAGE3_PCT":
            return sum(values[:3]) / 3.0 * 100.0
        if operation == "PRODUCT3_PCT":
            return values[0] * values[1] * values[2] * 100.0
        if operation == "MIN2":
            return min(values[0], values[1])
        if operation == "MULTIPLY_CONSTANT":
            if parameter is None:
                raise ValueError("MULTIPLY_CONSTANT requires parameter.")
            return values[0] * parameter
        if operation == "DIVIDE_CONSTANT":
            if parameter in (None, 0):
                raise ValueError("DIVIDE_CONSTANT requires a nonzero parameter.")
            return values[0] / parameter
        if operation == "SUM_RATIO_PCT":
            return _safe_divide(values[0] + values[1], values[2]) * 100.0
        if operation == "ROSI_PCT":
            return _safe_divide(values[0] - values[1], values[1]) * 100.0
        raise ValueError(f"Unsupported MRV operation: {operation}")

    def _ghg_from_energy_factors(
        self, values: list[float], config: Mapping[str, Any], scenario: Mapping[str, Any]
    ) -> float:
        factor_set_id = _text(scenario.get("emission_factor_set_id")) or self.default_factor_set_id
        electricity_code = _text(config.get("electricity_factor"))
        diesel_code = _text(config.get("diesel_factor"))
        divisor = _float_or_none(config.get("output_divisor"))
        if not factor_set_id or not electricity_code or not diesel_code or divisor in (None, 0):
            raise ValueError("Factor set, electricity factor, diesel factor, and nonzero output divisor are required")
        timestamp = pd.to_datetime(scenario.get("evaluation_timestamp"), errors="coerce")
        electricity_factor = self._resolve_analytical_factor(
            factor_set_id, electricity_code, timestamp,
            expected_unit="kgCO2e/kWh", required_role="active analytical",
        )
        diesel_factor = self._resolve_analytical_factor(
            factor_set_id, diesel_code, timestamp,
            expected_unit="kgCO2e/kWh", required_role="active analytical",
        )
        return (values[0] * electricity_factor + values[1] * diesel_factor) / float(divisor)

    def _multiply_factor(
        self, values: list[float], config: Mapping[str, Any], scenario: Mapping[str, Any]
    ) -> float:
        factor_set_id = _text(scenario.get("emission_factor_set_id")) or self.default_factor_set_id
        factor_code = _text(config.get("factor_code"))
        if not factor_set_id or not factor_code:
            raise ValueError("MULTIPLY_FACTOR requires factor_set_id and factor_code")
        timestamp = pd.to_datetime(scenario.get("evaluation_timestamp"), errors="coerce")
        factor = self._resolve_analytical_factor(
            factor_set_id, factor_code, timestamp,
            expected_unit="tCO2e/tkm",
            required_role=("transport-scope", "active analytical"),
            required_scope="transport",
        )
        return values[0] * factor

    def _resolve_analytical_factor(
        self, factor_set_id: str, factor_code: str, timestamp: pd.Timestamp,
        *, expected_factor_type: str = "EMISSION",
        expected_unit: str = "kgCO2e/kWh",
        required_role: str | tuple[str, ...] = "active analytical",
        required_scope: str | None = None,
    ) -> float:
        diagnostic = self.diagnose_analytical_factor(
            factor_set_id,
            factor_code,
            timestamp,
            expected_factor_type=expected_factor_type,
            expected_unit=expected_unit,
            required_role=required_role,
            required_scope=required_scope,
        )
        conditions = diagnostic["conditions"]
        if not conditions["registry_available"]:
            raise ValueError(f"Factor register is unavailable for {factor_code}")
        if not conditions["exactly_one_record"]:
            raise ValueError(
                f"Expected exactly one factor {factor_set_id}/{factor_code}; "
                f"found {diagnostic['matching_record_count']}"
            )
        if not conditions["factor_type_matches"]:
            raise ValueError(f"Factor {factor_code} must use factor type {expected_factor_type}")
        if not conditions["approval_status_approved"]:
            raise ValueError(f"Factor {factor_code} is not approved")
        if not conditions["analytical_role_allowed"]:
            raise ValueError(f"Factor {factor_code} is not authorized for analytical calculation")
        if not conditions["unit_matches"]:
            raise ValueError(f"Factor {factor_code} must use {expected_unit}")
        if not conditions["scope_matches"]:
            raise ValueError(f"Factor {factor_code} does not have the required {required_scope} scope")
        if not conditions["evaluation_timestamp_present"]:
            raise ValueError("Scenario timestamp is required for factor validity validation")
        if not conditions["valid_from_satisfied"] or not conditions["valid_to_satisfied"]:
            raise ValueError(f"Factor {factor_code} is not valid at {timestamp.date()}")
        if not conditions["finite_value"]:
            raise ValueError(f"Factor {factor_code} has no finite value")
        return float(diagnostic["record"]["value"])

    def diagnose_analytical_factor(
        self,
        factor_set_id: str,
        factor_code: str,
        timestamp: pd.Timestamp,
        *,
        expected_factor_type: str = "EMISSION",
        expected_unit: str = "kgCO2e/kWh",
        required_role: str | tuple[str, ...] = "active analytical",
        required_scope: str | None = None,
    ) -> dict[str, Any]:
        """Return the exact record and booleans used by factor authorization."""
        registry = self.factor_register
        registry_available = not registry.empty
        if registry_available:
            set_matches = registry[
                registry["factor_set_id"].astype(str).str.strip() == factor_set_id
            ]
            matches = set_matches[
                set_matches["factor_code"].astype(str).str.strip() == factor_code
            ]
        else:
            set_matches = registry
            matches = registry

        exactly_one = len(matches) == 1
        factor = matches.iloc[0] if exactly_one else None
        accepted_roles = (required_role,) if isinstance(required_role, str) else required_role
        evaluation_timestamp = pd.to_datetime(timestamp, errors="coerce")
        record = factor.to_dict() if factor is not None else None

        if factor is None:
            factor_type_matches = False
            approval_status_approved = False
            analytical_role_allowed = False
            unit_matches = False
            scope_matches = False
            valid_from_satisfied = False
            valid_to_satisfied = False
            finite_value = False
        else:
            factor_type_matches = (
                _text(factor.get("factor_type")).upper() == expected_factor_type.upper()
            )
            approval_status_approved = (
                _text(factor.get("approval_status")).lower() == "approved"
            )
            role = _text(factor.get("analytical_role")).lower()
            analytical_role_allowed = any(
                accepted_role.lower() in role for accepted_role in accepted_roles
            )
            unit_matches = _text(factor.get("unit")) == expected_unit
            scope_matches = required_scope is None or (
                required_scope.lower() in _text(factor.get("scope")).lower()
            )
            valid_from = pd.to_datetime(factor.get("valid_from"), errors="coerce")
            valid_to = pd.to_datetime(factor.get("valid_to"), errors="coerce")
            timestamp_present = not pd.isna(evaluation_timestamp)
            valid_from_satisfied = timestamp_present and (
                pd.isna(valid_from) or evaluation_timestamp >= valid_from
            )
            valid_to_satisfied = timestamp_present and (
                pd.isna(valid_to) or evaluation_timestamp <= valid_to
            )
            finite_value = _float_or_none(factor.get("value")) is not None

        conditions = {
            "registry_available": bool(registry_available),
            "factor_set_id_matches": bool(len(set_matches) > 0),
            "factor_code_matches_within_set": bool(len(matches) > 0),
            "exactly_one_record": bool(exactly_one),
            "factor_type_matches": bool(factor_type_matches),
            "approval_status_approved": bool(approval_status_approved),
            "analytical_role_allowed": bool(analytical_role_allowed),
            "unit_matches": bool(unit_matches),
            "scope_matches": bool(scope_matches),
            "evaluation_timestamp_present": bool(not pd.isna(evaluation_timestamp)),
            "valid_from_satisfied": bool(valid_from_satisfied),
            "valid_to_satisfied": bool(valid_to_satisfied),
            "finite_value": bool(finite_value),
        }
        return {
            "requested": {
                "factor_set_id": factor_set_id,
                "factor_code": factor_code,
                "evaluation_timestamp": evaluation_timestamp,
            },
            "expected": {
                "factor_type": expected_factor_type,
                "unit": expected_unit,
                "accepted_analytical_roles": list(accepted_roles),
                "required_scope": required_scope,
            },
            "matching_record_count": len(matches),
            "record": record,
            "conditions": conditions,
            "authorized": all(conditions.values()),
        }

    @staticmethod
    def _relationship_check(
        state: Mapping[str, Mapping[str, Any]],
        numerator: str,
        denominator: str,
        message: str,
        issue: Callable[..., None],
    ) -> None:
        if numerator not in state or denominator not in state:
            return
        left = _float_or_none(state[numerator].get("value"))
        right = _float_or_none(state[denominator].get("value"))
        if left is None or right is None:
            return
        if left > right + 1e-9:
            issue("QA_RELATIONSHIP", "Critical", "FAIL", message, variable=numerator)

    @staticmethod
    def _source_system(rule_level: str, source_module: Any) -> str:
        if rule_level == "L1":
            module = _slug(_text(source_module))
            return f"direct_model_output_{module}" if module else "direct_mrv_input"
        return RULE_SOURCE_SYSTEM.get(rule_level, "scenario_completion_engine")


# ----------------------------------------------------------------------
# Helper functions
# ----------------------------------------------------------------------
def _clean_frame(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    result.columns = [str(c).strip() for c in result.columns]
    result = result.dropna(how="all")
    return result


def _text(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return str(value).strip()


def _required_text(value: Any, field: str) -> str:
    text = _text(value)
    if not text:
        raise ValueError(f"Required scenario field {field!r} is blank.")
    return text


def _float_or_none(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        converted = float(value)
    except (TypeError, ValueError):
        return None
    return converted if math.isfinite(converted) else None


def _yes(value: Any) -> bool:
    return _text(value).lower() in {"yes", "true", "1", "y", "approved", "active"}


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def _selected_strategies(scenario: pd.Series) -> list[str]:
    values = [
        _text(scenario.get("primary_strategy")),
        _text(scenario.get("secondary_strategy_1")),
        _text(scenario.get("secondary_strategy_2")),
    ]
    return list(dict.fromkeys(v for v in values if v))


def _is_activity_dependent_primary(meta: Mapping[str, Any]) -> bool:
    """Return whether a primary quantity scales with the configured activity driver.

    The decision is based on the common-MRV semantic contract, not a case or
    scenario. Ratios, indices, headcounts, service counters, capital investment,
    and benefit assumptions are deliberately excluded.
    """
    if _text(meta.get("data_role")) != "PRIMARY_MRV":
        return False
    return _yes(meta.get("l3_scaling_eligible"))


def _is_physical_activity_quantity(meta: Mapping[str, Any]) -> bool:
    """Conservatively recognize scalable physical flows for scoped L3 rules.

    Counts, time, currency, percentages and indices need an exact declaration;
    physical resource/output units may use an explicit group-scope L3 permission.
    """
    if _text(meta.get("data_role")) != "PRIMARY_MRV":
        return False
    unit = _text(meta.get("canonical_unit")).lower().replace("³", "3")
    return unit in {"fu", "kwh", "m3", "t", "tkm", "kg", "l", "tonne"}


def _expand_rule_expression(expression: Any) -> set[str]:
    return parse_completion_levels(expression)


def parse_completion_levels(expression: Any) -> set[str]:
    """Parse explicit levels and inclusive ranges without substring matching."""
    text = _text(expression).upper().replace(" ", "")
    if not text:
        return set()
    result: set[str] = set()
    for token in re.split(r"[,/;]", text):
        match = re.fullmatch(r"L([1-6])(?:-L?([1-6]))?", token)
        if not match:
            raise ValueError(f"Unsupported completion-level expression: {expression!r}")
        first = int(match.group(1))
        last = int(match.group(2)) if match.group(2) else first
        for number in range(min(first, last), max(first, last) + 1):
            result.add(f"L{number}")
    return result


def _rule_sources(rule: Mapping[str, Any]) -> list[str]:
    sources = []
    for column in ("source_1", "source_2", "source_3"):
        value = _text(rule.get(column))
        if value:
            sources.append(value)
    return sources


def _safe_divide(numerator: float, denominator: float) -> float:
    if denominator == 0:
        raise ZeroDivisionError("MRV rule denominator is zero.")
    return numerator / denominator
