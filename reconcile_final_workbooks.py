"""Build the authoritative reconciled MRV fixtures from the versioned FINAL books."""
from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("tests/fixtures/mrv_final")
ACTIVITY_VARIABLES = (
    "electricity_kwh",
    "diesel_kwh",
    "water_withdrawn_m3",
    "waste_generated_t",
    "waste_recovered_t",
    "material_total_t",
    "material_circular_t",
    "transport_work_tkm",
    "operating_cost_eur",
)


def _headers(ws) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in ws[4] if cell.value is not None}


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    for source in ws[source_row]:
        target = ws.cell(target_row, source.column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _remove_unvalidated_sd_evidence(wb) -> None:
    ws = wb["02_DIRECT_MRV_INPUT"]
    columns = _headers(ws)
    scenario_col = columns["scenario_code"]
    variable_col = columns["variable_name"]
    for row in range(ws.max_row, 4, -1):
        scenario = str(ws.cell(row, scenario_col).value or "")
        variable = str(ws.cell(row, variable_col).value or "")
        if scenario.startswith("SD_") and variable == "mrv_points_active":
            ws.delete_rows(row)

    expected = wb["11_EXPECTED_CASE_MRV"]
    expected_columns = _headers(expected)
    reconciled_values = {
        "mrv_points_active": 140.0,
        "mrv_points_active_valid": 140.0,
        "mrv_coverage": 70.0,
    }
    for row in range(5, expected.max_row + 1):
        scenario = str(expected.cell(row, expected_columns["scenario_code"]).value or "")
        variable = str(expected.cell(row, expected_columns["variable_name"]).value or "")
        if scenario.startswith("SD_") and variable in reconciled_values:
            expected.cell(row, expected_columns["expected_value"], reconciled_values[variable])
            expected.cell(
                row,
                expected_columns["comment"],
                "Regression expectation after deactivating the unvalidated SD MRV index bridge.",
            )


def _add_des_activity_overrides(wb, *, active: str) -> None:
    ws = wb["08_VARIABLE_OVERRIDES"]
    columns = _headers(ws)
    existing = {
        (str(ws.cell(row, columns["strategy_code"]).value or ""),
         str(ws.cell(row, columns["variable_name"]).value or ""))
        for row in range(5, ws.max_row + 1)
    }
    source_row = max(5, ws.max_row)
    for variable in ACTIVITY_VARIABLES:
        if ("LOGISTICS_REDESIGN", variable) in existing:
            continue
        row = ws.max_row + 1
        _copy_row_style(ws, source_row, row)
        values = {
            "strategy_code": "LOGISTICS_REDESIGN",
            "variable_name": variable,
            "influence_status": "Activity-dependent scaling explicitly configured",
            "permitted_rules": "L1,L3,L4,L5,L6",
            "priority": 100,
            "scientific_justification": (
                "Scale from the reference intensity only when baseline scaling is enabled "
                "and the configured scenario driver is available."
            ),
            "active": active,
        }
        for name, value in values.items():
            ws.cell(row, columns[name], value)


def reconcile(source: Path, destination: Path, *, cuba: bool) -> None:
    wb = load_workbook(source)
    if cuba:
        _remove_unvalidated_sd_evidence(wb)
    _add_des_activity_overrides(wb, active="Yes" if cuba else "No")
    destination.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destination)


def main() -> None:
    reconcile(
        ROOT / "SustainSCM_Cuba_MRV_Scenario_Completion_FINAL.xlsx",
        ROOT / "SustainSCM_Cuba_MRV_Scenario_Completion_FINAL_RECONCILED.xlsx",
        cuba=True,
    )
    reconcile(
        ROOT / "SustainSCM_MRV_Causal_Completion_Template_FINAL.xlsx",
        ROOT / "SustainSCM_MRV_Causal_Completion_Template_FINAL_RECONCILED.xlsx",
        cuba=False,
    )


if __name__ == "__main__":
    main()
